import tkinter as tk
from tkinter import filedialog, Menu
from tkinter import Canvas
import json
from pdf2image import convert_from_path
from PIL import Image, ImageTk
import openpyxl
from tkinter import Toplevel, ALL
import re
import os

def split_excel_cell(excel_cell):
    match = re.match(r'([A-Za-z]+)(\d+)', excel_cell)
    if match:
        column, row = match.groups()
        return column, int(row)
    else:
        return None, None

class DataEntryApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Data Entry App")
        self.master.geometry("1000x800")
        self.create_widgets()
        self.current_page = 0
        self.current_roi = 0
        self.template_roi = None  # Add this line to initialize the template_roi attribute
        self.original_calibration_roi = None  # Add this line to initialize the original_calibration_roi attribute
        self.previous_mouse_position = (0, 0)
        self.page_number = None
        self.entered_data = {}
        self.last_printed_key = None
        self.training_data_on = False

    def close_window(self):
        self.entered_data = {}
        self.window.destroy()

    def main(self):
        self.window.protocol("WM_DELETE_WINDOW", self.close_window)
        self.window.mainloop()

    def turn_on_training_data_no_prompt(self):
        if self.training_data_out_loc is None:
            messagebox.showerror("Error", "Please set the directory location first through the file menu.")
        else:
            self.training_data_on = True

    def turn_on_training_data(self):
        self.training_data_on = True
        self.training_data_out_loc = filedialog.askdirectory()
        self.training_button.config(text='On', fg='green', font=('Helvetica', '16', 'bold'))

    def turn_off_training_data(self):
        self.training_data_on = False

    def toggle_training_data(self):
        if self.training_data_on:
            self.turn_off_training_data()
            self.training_button.config(text='Off', fg='red', font=('Helvetica', '16', 'bold'))
        else:
            self.turn_on_training_data_no_prompt()
            self.training_button.config(text='On', fg='green', font=('Helvetica', '16', 'bold'))


    def create_widgets(self):
        # Create menu
        self.menu = Menu(self.master)
        self.master.config(menu=self.menu)
        self.file_menu = Menu(self.menu, tearoff=0)
        self.menu.add_cascade(label="File", menu=self.file_menu)
        self.file_menu.add_command(label="Upload PDF", command=self.load_pdf_pages)
        self.file_menu.add_command(label="Upload JSON", command=self.load_template_roi)
        self.file_menu.add_command(label="Save Progress", command=self.save_data)
        self.file_menu.add_command(label="Write to Excel", command=self.write_to_excel_window)
        self.file_menu.add_command(label="Adjust The Template", command=self.open_calibration_window)
        # Create Training Data submenu
        self.training_data_menu = Menu(self.file_menu, tearoff=0)
        self.file_menu.add_cascade(label="Training Data", menu=self.training_data_menu)
        self.training_data_menu.add_command(label="Turn On (Select Out-Loc)", command=self.turn_on_training_data)
        self.training_data_menu.add_command(label="Turn Off", command=self.turn_off_training_data)
        #Training Data Button
        self.training_button = tk.Button(self.master, text="On/Off", command=self.toggle_training_data)
        self.training_button.place(relx=0.70, rely=0.85, anchor='n')
        self.trainbutton_label = tk.Label(self.master, text="Training Data", font=("Helvetica", 12, "bold"))
        self.trainbutton_label.place(relx=0.7, rely=0.83, anchor='s')

        #Create Page Number Label
        self.page_num_label = tk.Label(self.master, font=("Helvetica", 14, "bold") )
        self.page_num_label.place(relx=0.35, rely=0.1, anchor='s')

        # Create ROI name label
        self.roi_name_label = tk.Label(self.master, font=("Helvetica", 14, "bold"))
        self.roi_name_label.place(relx=0.65, rely=0.1, anchor='s')

        self.j_son_label = tk.Label(self.master)
        self.j_son_label.place(relx=0.15, rely=0.05, anchor='s')

        self.pdf_file_label = tk.Label(self.master)
        self.pdf_file_label.place(relx=0.85, rely=0.05, anchor='s')

        # Create a canvas for the ROI display
        self.roi_display_canvas = Canvas(self.master, width=700, height=375, bg="white", bd=2, relief="groove")
        self.roi_display_canvas.place(relx=0.5, rely=0.35, anchor='center')

        # Create Upper ROI display region
        self.roi_display_prev = tk.Label(self.roi_display_canvas)
        self.roi_display_prev.place(relx=0.5, rely=0.22, anchor='center')

        # Create ROI display region
        self.roi_display = tk.Label(self.roi_display_canvas, relief="groove")
        self.roi_display.place(relx=0.5, rely=0.5, anchor='center')

        # Create Bottom ROI display region
        self.roi_display_next = tk.Label(self.roi_display_canvas)
        self.roi_display_next.place(relx=0.5, rely=0.78, anchor='center')

        # Create text entry widget
        self.text_entry = tk.Entry(self.master)
        self.text_entry.place(relx=0.5, rely=0.7, anchor='n')
        # Bind Enter Key
        self.text_entry.bind('<Return>', self.on_enter_key_pressed)

        # Create ROI navigation label and buttons
        self.roi_label = tk.Label(self.master, text="ROI")
        self.roi_label.place(relx=0.5, rely=0.75, anchor='n')

        self.roi_next_button = tk.Button(self.master, text="NEXT>>>", command=self.next_roi)
        self.roi_next_button.place(relx=0.535, rely=0.80, anchor='n')

        self.roi_back_button = tk.Button(self.master, text="<<<BACK", command=self.prev_roi)
        self.roi_back_button.place(relx=0.465, rely=0.80, anchor='n')

        # Create page navigation label and buttons
        self.page_label = tk.Label(self.master, text="Page")
        self.page_label.place(relx=0.50, rely=0.85, anchor='n')

        self.page_next_button = tk.Button(self.master, text="NEXT>>>", command=self.next_page)
        self.page_next_button.place(relx=0.535, rely=0.90, anchor='n')

        self.page_back_button = tk.Button(self.master, text="<<<BACK", command=self.prev_page)
        self.page_back_button.place(relx=0.465, rely=0.90, anchor='n')

    def load_template_roi(self):
        # Prompt the user to select a JSON file
        file_path = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])

        if not file_path:
            return

        # Load ROI template from the .json file
        with open(file_path, 'r') as json_file:
            new_template_roi = json.load(json_file)

        self.template_roi = new_template_roi  # Update the template ROI without applying previous calibration
        self.find_calibration_roi()  # Find the "Calibrate" ROI
        self.original_calibration_roi = self.calibration_roi.copy()  # Reset the original calibration ROI
        self.setup_page_iteration()

        # Get the JSON file name without the path
        json_file_name = os.path.basename(file_path)
        
        # Extract the prefix (should be either 50 or 25) from the filename
        self.json_prefix = int(json_file_name.split('-')[0])
        self.json_type_suffix = json_file_name.split('-')[2].split('_')[0]

        self.j_son_label.config(text=json_file_name)

        self.open_calibration_window()


    def load_pdf_pages(self):
        # Prompt the user to select a PDF file
        file_path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
        pdf_file_name = os.path.basename(file_path)
        self.pdf_file_label.config(text=pdf_file_name)

        if not file_path:
            return

        # Convert each page of the PDF to a PNG image
        self.pdf_images = convert_from_path(file_path, fmt="png", dpi=300)

        # Set the current page to the first page
        self.current_page = 0
        self.display_roi()
        if self.template_roi:  # Check if the JSON file has been loaded
            self.setup_page_iteration()
            self.open_calibration_window()

    def open_calibration_window(self):
        if not self.template_roi or not self.pdf_images:
            return

        self.original_calibration_roi = None
        self.find_calibration_roi()

        self.calibration_window = Toplevel(self.master)
        self.calibration_window.title("Calibration")


        if self.current_page == 0:
            self.page_image = self.pdf_images[1]
        elif self.current_page > 0:
            # Display the second page of the PDF
            self.page_image = self.pdf_images[self.current_page]

        # Create a canvas for the calibration display
        self.calibration_display_canvas = Canvas(self.calibration_window, width=500, height=500, bg="white")
        self.calibration_display_canvas.pack(pady=10)

        self.display_calibration_roi()

        # Add a label with instructions
        self.instructions_label = tk.Label(self.calibration_window, text="Please reposition the red dotted-line box so that the '1' is centered")
        self.instructions_label.pack()

        # Add a Calibrate button
        self.calibrate_button = tk.Button(self.calibration_window, text="Calibrate", command=self.calibrate)
        self.calibrate_button.pack(pady=10)
        # Set the value of the original_calibration_roi attribute when the red box is first drawn
        if self.original_calibration_roi is None:
            self.original_calibration_roi = self.calibration_roi.copy()


    def find_calibration_roi(self):
        for roi in self.template_roi:
            if roi['name'] == 'Calibrate':
                self.calibration_roi = roi
                break
                
    def display_calibration_roi(self):
        if not self.template_roi or not self.pdf_images:
            return

        if self.current_page == 0:
            self.page_image = self.pdf_images[1]
        elif self.current_page > 0:
            # Display the second page of the PDF
            self.page_image = self.pdf_images[self.current_page]

        x1, y1, x2, y2 = self.calibration_roi['x1'], self.calibration_roi['y1'], self.calibration_roi['x2'], self.calibration_roi['y2']

        # Calculate the coordinates of the 500px by 500px region surrounding the ROI
        region_x1 = max(x1 - 250, 0)
        region_y1 = max(y1 - 250, 0)
        region_x2 = region_x1 + 500
        region_y2 = region_y1 + 500

        # Crop the 500px by 500px region from the PDF page image
        cropped_image = self.page_image.crop((region_x1, region_y1, region_x2, region_y2))
        self.region_image = ImageTk.PhotoImage(cropped_image)

        # Display the 500px by 500px region on the canvas
        self.calibration_display_canvas.create_image(250, 250, image=self.region_image)

        # Display the "Calibrate" ROI as a red dotted-line box overlay
        roi_x1, roi_y1 = x1 - region_x1, y1 - region_y1
        roi_x2, roi_y2 = x2 - region_x1, y2 - region_y1
        self.calibration_display_canvas.create_rectangle(roi_x1, roi_y1, roi_x2, roi_y2, outline="red", dash=(3, 3), width=2)

        # Bind mouse movement events to the adjust_roi method
        self.calibration_display_canvas.bind("<B1-Motion>", self.adjust_roi)
        self.calibration_display_canvas.bind("<Button-1>", self.set_previous_mouse_position)


    def calibrate(self):
        # Calculate the difference between the original and new ROI positions
        x_diff = self.calibration_roi['x1'] - self.original_calibration_roi['x1']
        y_diff = self.calibration_roi['y1'] - self.original_calibration_roi['y1']

        # Create a new JSON with the updated ROI positions
        new_template_roi = []
        for roi in self.template_roi:
            if roi['name'] != "Calibrate":
                new_roi = roi.copy()
                new_roi['x1'] += x_diff
                new_roi['y1'] += y_diff
                new_roi['x2'] += x_diff
                new_roi['y2'] += y_diff
                new_template_roi.append(new_roi)

        # Replace the current JSON with the new JSON
        self.template_roi = new_template_roi

        # Close the calibration window
        self.calibration_window.destroy()

        if self.page_number == None:
            self.next_page()
            self.page_num_label.config(text='Enter Page Number')
        elif self.page_number >= 1:
            self.display_roi()


    def adjust_roi(self, event):
        # Scaling factor to make the red box movement less sensitive
        scaling_factor = 0.3

        # Calculate the difference between the current and previous mouse positions
        dx = (event.x - self.previous_mouse_position[0]) * scaling_factor
        dy = (event.y - self.previous_mouse_position[1]) * scaling_factor

        # Update the previous mouse position
        self.previous_mouse_position = (event.x, event.y)

        # Update the ROI position based on the differences in mouse position
        self.calibration_roi['x1'] += dx
        self.calibration_roi['y1'] += dy
        self.calibration_roi['x2'] += dx
        self.calibration_roi['y2'] += dy

        # Redraw the calibration display
        self.calibration_display_canvas.delete(ALL)
        self.display_calibration_roi()

    def set_previous_mouse_position(self, event):
        self.previous_mouse_position = (event.x, event.y)


    def setup_page_iteration(self):
        self.current_page = 0
        self.current_roi = 0
        self.display_roi()



    def display_roi(self):
        if not self.template_roi or not self.pdf_images:
            return

        page_image = self.pdf_images[self.current_page]

        rois = {
            'current': self.template_roi[self.current_roi] if len(self.template_roi) > self.current_roi else None,
            'prev': self.template_roi[self.current_roi - 1] if self.current_roi - 1 >= 0 else None,
            'next': self.template_roi[self.current_roi + 1] if self.current_roi + 1 < len(self.template_roi) else None
        }

        for key, roi in rois.items():
            if roi is None:
                continue

            x1, y1, x2, y2 = roi['x1'], roi['y1'], roi['x2'], roi['y2']
            cropped_image = page_image.crop((x1, y1, x2, y2))
            roi_image = ImageTk.PhotoImage(cropped_image)

            if key == 'current':
                self.roi_image = roi_image
                self.original_roi_image = cropped_image
                self.roi_display.config(image=self.roi_image)
                self.roi_name_label.config(text=roi['name'])
            elif key == 'prev':
                self.roi_image_prev = roi_image
                self.original_roi_image_prev = cropped_image
                self.roi_display_prev.config(image=self.roi_image_prev)
                # Update the ROI name label if necessary
            elif key == 'next':
                self.roi_image_next = roi_image
                self.original_roi_image_next = cropped_image
                self.roi_display_next.config(image=self.roi_image_next)
                # Update the ROI name label if necessary

    def save_training_data(self):
        # First, save the image
        image_path = f'{self.training_data_out_loc}/{self.current_roi_name}_{self.page_number}.png'
        self.original_roi_image.save(image_path)

        # Now, save the corresponding text
        text_path = f'{self.training_data_out_loc}/{self.current_roi_name}_{self.page_number}.txt'
        with open(text_path, 'w') as text_file:
            text_file.write(self.entered_value)

    def on_enter_key_pressed(self, event):
        self.current_roi_name = self.template_roi[self.current_roi]['name']
        self.entered_value = self.text_entry.get()
        self.detected_triggers = self.extract_triggers(self.entered_value)
        self.roi_index = self.calculate_roi_index()
        
        if self.current_roi_name == "Page_Count":
            self.process_page_count()
        elif self.text_entry.get() == "":
            pass
        else:
            self.process_non_empty_entry()

            if not self.detected_triggers and self.entered_value.strip() != '' and self.training_data_on:
                self.save_training_data()

            
        self.text_entry.delete(0, 'end')
        
        if self.entered_data:
            last_key, last_data = list(self.entered_data.items())[-1]
            # Only print if the last key is different from the last printed key
            if last_key != self.last_printed_key:
                print(json.dumps((last_key, last_data), indent=2))
                self.last_printed_key = last_key  # Update the last printed key
            
        self.next_roi()
        
          
    def extract_triggers(self, entered_value):
        if self.json_type_suffix == 'PDPIR' or self.json_type_suffix == 'HWDP':
            triggers = [
                "MW", "MBD", "DB", "R", "EMI", "DT", "DS", "ODAM", "SB", "OR", 
                "OT", "HB", "LB", "MT", "MS", "LP", "BNT", "DAM", "MOD", "OTHER"
            ]
            return [word for word in entered_value.split() if word in triggers]
        elif self.json_type_suffix == 'TUBING':
            triggers = [
                "MW", "RW", "DB", "DP", "HB", "PIT", "MASH", "SC", "SCR", "GOU", 
                "TC", "BNT", "DBR", "NODRIFT", "EMI"
            ]
            return [word for word in entered_value.split() if word in triggers]
            


    def calculate_roi_index(self):
        roi_index = 0
        for roi in self.template_roi:
            # Check if the roi name starts with the current_roi_name (before the '_')
            if roi['name'].startswith(self.current_roi_name.split('_')[0]):
                # If it's the same as the current roi, break the loop
                if roi == self.template_roi[self.current_roi]:
                    break
                else:
                    # If it's not the same, but starts with the same string, increment the index
                    roi_index += 1
        return roi_index

    def process_page_count(self):
        self.page_number = int(self.text_entry.get())
        print(f"Page Number set to: {self.page_number}")
        self.page_num_label.config(text='Page Count: ' + str(self.current_page))
        
    def process_non_empty_entry(self):   
        self.process_non_trigger_value()
        self.process_triggers()

    def process_non_trigger_value(self):
        non_trigger_value = ' '.join([v for v in self.entered_value.split() if v not in self.detected_triggers])
        if non_trigger_value and ('excel_cell' in self.template_roi[self.current_roi] or 'excel_cell1' in self.template_roi[self.current_roi]):
            new_excel_cell = None
            new_excel_cell1 = None
            if 'excel_cell' in self.template_roi[self.current_roi]:
                new_excel_cell = self.calculate_new_excel_cell(self.template_roi[self.current_roi]['excel_cell'])
            if 'excel_cell1' in self.template_roi[self.current_roi]:
                new_excel_cell1 = self.calculate_new_excel_cell(self.template_roi[self.current_roi]['excel_cell1'])
            roi_data_key = f"{self.current_roi_name}_{self.page_number}"
            self.entered_data[roi_data_key] = {"value": non_trigger_value, "excel_cell": new_excel_cell, "excel_cell1": new_excel_cell1}

    def process_triggers(self):
        yellow_trigs = [
            "bent_tube_BENT", "bent_tube_BENT DBR", "box_connclass_DB", "pin_connclass_DP"
        ]
        red_trigs = [
            "gouge_slipcut_SC", "gouge_slipcut_GOU", "gouge_slipcut_TC", 
            "tube_condition_MW", "tube_condition_RW", "tube_condition_Pit", 
            "tube_condition_Mashed", "fl_drift_NO", "emi_tube_EMI"
        ]
        
        final_class_key = None
        for detected_trigger in self.detected_triggers:
            trigger_keys = self.trigger_to_key(detected_trigger)
            if not isinstance(trigger_keys, list):
                trigger_keys = [trigger_keys]

            for trigger_key in trigger_keys:
                self.update_entered_data(trigger_key)
            
            # determine which final class to use
            if trigger_key in red_trigs and final_class_key != "final_class_Scrap":
                final_class_key = "final_class_Scrap"
            elif trigger_key in yellow_trigs and final_class_key is None:
                final_class_key = "final_class_Repairable"
        
        # update entered data with the final class item, if one was found
        if final_class_key:
            self.update_entered_data(final_class_key, final_class=True)

    def update_entered_data(self, trigger_key, final_class=False):
        if final_class:
            # directly use the provided final class key
            new_excel_cell = self.calculate_new_excel_cell("N10")
            additional_item = {trigger_key: new_excel_cell}
            roi_data_key = f"{self.current_roi_name}_{self.page_number}"
            if roi_data_key in self.entered_data:
                self.entered_data[roi_data_key].update(additional_item)
            else:
                self.entered_data[roi_data_key] = additional_item
        elif trigger_key in self.template_roi[self.current_roi]:
            new_excel_cell = self.calculate_new_excel_cell(self.template_roi[self.current_roi][trigger_key])
            additional_item = {trigger_key: new_excel_cell}
            roi_data_key = f"{self.current_roi_name}_{self.page_number}"
            if roi_data_key in self.entered_data:
                self.entered_data[roi_data_key].update(additional_item)
            else:
                self.entered_data[roi_data_key] = additional_item

    def calculate_new_excel_cell(self, excel_cell):
        column, initial_row = split_excel_cell(excel_cell)  # Not using trigger_key here
        
        # Use the json_prefix (either 50 or 25) to calculate the new row
        new_row = initial_row + (self.page_number - 1) * self.json_prefix + self.roi_index
        new_excel_cell = f"{column}{new_row}"
        
        return new_excel_cell



    def trigger_to_key(self, trigger):

        if self.json_type_suffix == 'PDPIR' or self.json_type_suffix == 'HWDP':
            trigger_key_map = {
                "MW": "min_wall",
                "DB": ["dhb_box", "dhb_pin"],
                "R": ["box_reface", "pin_reface"],
                "DS": ["ds_box", "ds_pin"],
                "EMI": "emi_reject",
                "DT": ["dt_box", "dt_pin"],
                "OR": ["or_box", "or_pin"],
                "HB": ["hb_box", "hb_pin"],
                "MT": ["mintong_box", "mintong_pin"],
                "MS": ["minseal_box", "minseal_pin"],
                "ODAM": ["odam_box", "odam_pin"],
                "SB": "short_box",
                "LP": "long_pin",
                "BNT": "bent_tube",
                "DAM": "damaged_tube",
                "MOD": "min_od",
                "OTHER": "other_bp"
                # Add more trigger-key mappings here
            }

        elif self.json_type_suffix == 'TUBING':
            trigger_key_map = {
                "MW": "tube_condition_MW",
                "RW": "tube_condition_RW",
                "PIT": "tube_condition_Pit",
                "MASH": "tube_condition_Mashed",
                "SC": "gouge_slipcut_SC",
                "SCR": "gouge_slipcut_SC-R",
                "GOU": "gouge_slipcut_GOU",
                "TC": "gouge_slipcut_TC",
                "BNT": "bent_tube_BENT",
                "DBR": "bent_tube_BENT DBR",
                "NODRIFT": "fl_drift_NO",
                "EMI": "emi_tube_EMI",
                "DB": "box_connclass_DB",
                "DP": "pin_connclass_DP",
                "HB": "hb_placement_HB",
            }


        trigger_keys = trigger_key_map.get(trigger)
        if isinstance(trigger_keys, list):
            trigger_keys = [key for key in trigger_keys if key in self.template_roi[self.current_roi]]
        return trigger_keys

#--------------------------------------------------------------------
    def next_roi(self):
        if not hasattr(self, 'pdf_images') or not hasattr(self, 'template_roi'):
            return

        if self.text_entry.get() == "0" and current_roi_name == "Page_Count":
            self.next_page()
        elif self.current_roi == len(self.template_roi) - 1:  # Check if we are at the last ROI
            self.current_roi = 1  # Go back to the first ROI, skipping the "Calibrate" ROI
            self.page_number += 1  # Increment the page_number
            self.next_page()  # Load the next PDF page
        else:
            self.current_roi += 1
            self.display_roi()


    def prev_roi(self):
        if not hasattr(self, 'pdf_images') or not hasattr(self, 'template_roi'):
            return

        if self.current_roi > 0:
            self.current_roi -= 1
            self.display_roi()


    def next_page(self):
        if not hasattr(self, 'pdf_images') or not hasattr(self, 'template_roi'):
            return

        if self.current_page < len(self.pdf_images) - 1:
            self.current_page += 1
            self.current_roi = 0
            self.display_roi()
            self.page_num_label.config(text='Page Count: ' + str(self.current_page))

    def prev_page(self):
        if not hasattr(self, 'pdf_images') or not hasattr(self, 'template_roi'):
            return

        if self.current_page > 0:
            self.current_page -= 1
            self.current_roi = 0
            self.display_roi()
            self.page_num_label.config(text='Page Count: ' + str(self.current_page))


    def save_data(self, page_number):
        file_path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")])
        if not file_path:
            return

        with open(file_path, 'w') as json_file:
            for key, value in data.items():
                if key in self.template_roi:
                    # Extract column and initial row number from excel_cell
                    column, initial_row = split_excel_cell(self.template_roi[key]["excel_cell"])

                    if column and initial_row is not None:
                        # Calculate the new row number based on the current page_number
                        new_row = initial_row + (page_number - 1) * 50 + int(key.split('_')[-1]) - 1

                        # Update the excel_cell value with the new row number
                        new_excel_cell = f"{column}{new_row}"

                        # Save the updated ROI data in self.entered_data
                        self.entered_data[key] = {"value": value, "excel_cell": new_excel_cell}
                    else:
                        # If the excel_cell value is invalid, save the data without updating it
                        self.entered_data[key] = {"value": value, "excel_cell": self.template_roi[key]["excel_cell"]}

            json.dump(self.entered_data, json_file)



            
    def write_to_excel_window(self):
        # Prompt the user to select an Excel file
        excel_file = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")])
        if not excel_file:
            return

        # Store the selected Excel file as an attribute
        self.excel_file = excel_file

        # Create a new window with two buttons
        write_data_window = tk.Toplevel(self.master)
        use_current_data_button = tk.Button(write_data_window, text="Use Current Data", command=lambda: self.prepare_data(write_data_window, self.entered_data))
        upload_saved_data_button = tk.Button(write_data_window, text="Upload Saved Data", command=lambda: self.prepare_data_from_json(write_data_window, excel_file))

        use_current_data_button.pack()
        upload_saved_data_button.pack()



    def prepare_data(self, write_data_window, data):
        write_data_window.destroy()
        self.data_to_write = data
        self.show_write_data_button()

    def prepare_data_from_json(self, write_data_window, excel_file):
        write_data_window.destroy()
        json_file = filedialog.askopenfilename(filetypes=[("JSON Files", "*.json"), ("All Files", "*.*")])
        if not json_file:
            return

        with open(json_file, "r") as file:
            self.data_to_write = json.load(file)

        self.show_write_data_button()

    def show_write_data_button(self):
        write_data_button = tk.Button(self.master, text="WRITE DATA", command=self.write_prepared_data)
        write_data_button.grid(row=7, column=0, columnspan=2, pady=10)


    def write_prepared_data(self):
        if not self.entered_data:
            return

        # Open the selected Excel file
        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb.active

        # Create a set for storing row identifiers with final_class_Scrap
        scrap_rows = set()

        # Loop through all data items and identify rows that contain final_class_Scrap
        for key, data in self.entered_data.items():
            if 'final_class_Scrap' in data:
                # Extract row identifier from key
                row_identifier = key.split('_')[-2:]
                # Add row identifier to scrap_rows set
                scrap_rows.add('_'.join(row_identifier))

        # Write the data to the Excel file
        for key, data in self.entered_data.items():
            main_value = data.get("value")
            main_excel_cell = None

            # Determine which cell to write to based on the triggers detected, may cause issues if REFACE and SB, LP, MT are present at same time.
            if "box_reface" in data or "pin_reface" in data:
                main_excel_cell = data.get("excel_cell")
            elif "short_box" in data or "long_pin" in data or "mintong_box" in data or "mintong_pin" in data:
                main_excel_cell = data.get("excel_cell1")
            else:
                main_excel_cell = data.get("excel_cell")  # Default to "excel_cell"

            if main_value is not None and main_excel_cell is not None:
                ws[main_excel_cell] = main_value

            # Extract row identifier from key
            row_identifier = '_'.join(key.split('_')[-2:])

            for sub_key, cell_value in data.items():
                if sub_key not in ["value", "excel_cell", "excel_cell1"]:
                    # Handle 'final_class_Repairable' key for rows with 'final_class_Scrap'
                    if '_'.join(row_identifier) in scrap_rows and sub_key == 'final_class_Repairable':
                        continue  # Do not write 'final_class_Repairable' if row has 'final_class_Scrap'

                    value_to_write = ""  # Default value
                    if self.json_type_suffix == 'TUBING':
                        value_to_write = sub_key.split('_')[-1]  # Grabs the text after the last '_'
                    elif self.json_type_suffix == 'PDPIR' or self.json_type_suffix == 'HWDP':
                        value_to_write = "X"
                    # You can add more conditions here
                    ws[cell_value] = value_to_write

        # Save the Excel file
        wb.save(self.excel_file)




if __name__ == "__main__":
    root = tk.Tk()
    app = DataEntryApp(root)
    root.mainloop()
